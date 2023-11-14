import openpyxl

def save_row_to_excel(file_path, data):
    try:
        # Mở workbook từ tệp Excel đã tồn tại
        wb = openpyxl.load_workbook(file_path)
        # Lấy sheet mặc định
        sheet = wb.active
    except FileNotFoundError:
        # Nếu tệp không tồn tại, tạo mới workbook và sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Copy header từ file Excel mẫu (nếu có)
        sample_file_path = "sample.xlsx"
        try:
            sample_file = openpyxl.load_workbook(sample_file_path)
            sample_sheet = sample_file.active
            header_row = list(sample_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            sheet.append(header_row)
        except FileNotFoundError:
            print(f"Không tìm thấy tệp mẫu: {sample_file_path}")

    # Thêm dữ liệu từ dòng thứ 2 vào sheet
    sheet.append(data)

    # Lưu workbook vào một tệp Excel
    wb.save(file_path)

# Dữ liệu mẫu
data1 = ["a", "a", "a", "a", "a", "a", "a", "a", "a", "a", "a", "a", "a", "a", "a", "a", "a"]

# Lưu dữ liệu từ dòng thứ 2 vào tệp Excel
save_row_to_excel("data.xlsx", data1)
